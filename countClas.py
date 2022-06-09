from openpyxl import load_workbook,Workbook

wb = load_workbook("总名单.xlsx")
sh = wb["Sheet1"]

dic={}

for i in range(1,sh.max_row+1):
    if sh.cell(i,1).value not in dic:
        dic[sh.cell(i,1).value] = 1
    else:
        dic[sh.cell(i,1).value] += 1

wb1 = Workbook()
sh1 = wb1.active

i=0
for key,value in dic.items():
    i+=1
    sh1.cell(i,1).value = key
    sh1.cell(i,2).value = value

sh1.column_dimensions['A'].width = 45.0

wb1.save("总人数表.xlsx")