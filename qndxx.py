from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font,Alignment

class cl:
    grade = 0  #年级
    name = ''
    total = 0  #总计
    unfi = 0  #未完成
    per = 0  #完成率

    def __init__(self,n,t):
        self.name = n
        self.total = t

    def __lt__(self,others):
        return self.per < others.per

    def calPercent(self):
        self.per = 1 - (self.unfi/self.total)

    # def getPercent(self):
    #     return self.per

    def addUnfinished(self):
        self.unfi += 1

    def addTotal(self):
        self.total += 1

    # def getClassName(self):
    #     return self.name

class stu:
    name = ''
    clas = ''

    def __init__(self,n,c):
        self.name = n
        self.clas = c

    def __lt__(self,others):
        return self.clas < others.clas

def cnt(nam):
    for i in range(0,total_class):
        if arr[i].name == nam:
            arr[i].addUnfinished()
            return
    return

tarGrade=19
year=2022
issue=0

if issue == 0:
    wbtitle = str(tarGrade)+'级青年大学习'+str(year)+'年特辑'
else:
    wbtitle = str(tarGrade)+'级青年大学习'+str(year)+'年第'+str(issue)+'期'

weiwancheng='未完成名单'
wanchenglv='完成率'

wb0 = load_workbook("完成.xlsx")
# wb1 = load_workbook("未完成名单.xlsx")
wb2 = load_workbook("总名单.xlsx")
wb3 = load_workbook("总人数表.xlsx")
sh0 = wb0["报名列表"]
# sh1 = wb1["Sheet1"]
sh2 = wb2["Sheet1"]
sh3 = wb3["Sheet1"]
total_class=sh3.max_row
# total_unfinished_stu=sh1.max_row

fns_name = []
tot_stu = []
unfinished_stu = []

for i in range(2,sh0.max_row+1):
    fns_name.append(sh0.cell(i,2).value)

for i in range(1,sh2.max_row+1):
    tot_stu.append(stu(str(sh2.cell(i,2).value),str(sh2.cell(i,1).value)))

for x in tot_stu:
    if x.name not in fns_name:
        unfinished_stu.append(x)

unfinished_stu.sort()

wb1 = Workbook()
sh1 = wb1.active
sh1.cell(1,1).value = wbtitle + weiwancheng
sh1.merge_cells('A1:B1')
sh1.cell(2,1).value = '姓名'
sh1.cell(2,2).value = '班级'

for i in range(0,len(unfinished_stu)):
    sh1.cell(i+3,1).value = unfinished_stu[i].name
    sh1.cell(i+3,2).value = unfinished_stu[i].clas

style_1 = Font(u'宋体', size=14) 
style_2 = Alignment(horizontal='center', vertical='center')
for col in sh1["A:B"]:
    for cell in col:
        cell.font = style_1
        cell.alignment = style_2

sh1.column_dimensions['A'].width = 20.0
sh1.column_dimensions['B'].width = 45.0

wb1.save(wbtitle+weiwancheng+".xlsx")



arr=[]
for i in range(1,total_class+1):
    arr.append(cl(sh3.cell(i,1).value, sh3.cell(i,2).value))

for col in sh1.rows:
    cnt(col[1].value)

for i in range(0,total_class):
    arr[i].calPercent()

arr.sort(reverse=1)

wb4 = Workbook()
sh4 = wb4.active
sh4.cell(1,1).value = wbtitle + wanchenglv
sh4.merge_cells('A1:B1')
sh4.cell(2,1).value = '班级'
sh4.cell(2,2).value = '完成率'

for i in range(0,total_class):
    sh4.cell(i+3,1).value = arr[i].name
    sh4.cell(i+3,2).value = arr[i].per

for col in sh4["A:B"]:
    for cell in col:
        cell.font = style_1
        cell.alignment = style_2

for cell in sh4["B"]:
    cell.number_format = '0.00%'

sh4.column_dimensions['A'].width = 45.0
sh4.column_dimensions['B'].width = 15.0

for i in range(3,total_class+3):
    if sh4.cell(i,2).value < 0.7:
        for j in range(i,total_class+3):
            sh4.cell(j,1).font += Font(color="ff0000")
            sh4.cell(j,2).font += Font(color="ff0000")
        break

wb4.save(wbtitle+wanchenglv+".xlsx")
